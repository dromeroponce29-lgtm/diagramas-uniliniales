"""
Exportador XLSX profesional con openpyxl.
Genera un libro con varias hojas: Cubicación, Presupuesto, Normativa, Riesgos, Resumen.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# Estilos reutilizables
FILL_HEAD = PatternFill("solid", fgColor="0F2C4A")
FONT_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=10)
FONT_TOTAL = Font(name="Calibri", size=10, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
BORDER_THIN = Border(*[Side(style="thin", color="BBBBBB")] * 4)


def _aplicar_header(ws, fila, columnas):
    for i, col in enumerate(columnas, start=1):
        c = ws.cell(row=fila, column=i, value=col)
        c.fill = FILL_HEAD
        c.font = FONT_HEAD
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN


def _auto_width(ws, max_width=60):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for c in col_cells:
            if c.value is None:
                continue
            longitud = len(str(c.value))
            if longitud > max_len:
                max_len = longitud
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def exportar_xlsx(
    proyecto: dict,
    cubicacion: list[dict],
    normativa: list[dict],
    riesgos: list[dict],
    config: dict,
    output_path: Path,
) -> Path:
    wb = Workbook()

    # ===== Hoja Resumen =====
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "INFORME PROYECTO ELÉCTRICO"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0F2C4A")
    ws["A2"] = proyecto.get("nombre", "")
    ws["A2"].font = Font(name="Calibri", size=13, bold=True)
    ws["A3"] = f"Generado: {datetime.utcnow().strftime('%d-%m-%Y %H:%M UTC')}"
    ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="666666")

    datos = [
        ("Mandante", proyecto.get("mandante")),
        ("Ubicación", proyecto.get("ubicacion")),
        ("Tipo de instalación", proyecto.get("tipo")),
        ("Tensión (V)", proyecto.get("tension")),
        ("Potencia (kW)", proyecto.get("potencia")),
        ("Ambiente", proyecto.get("ambiente")),
        ("Método de instalación", proyecto.get("metodo_instalacion")),
        ("Normativa aplicable", proyecto.get("normas")),
    ]
    for i, (k, v) in enumerate(datos, start=5):
        ws.cell(row=i, column=1, value=k).font = FONT_TOTAL
        ws.cell(row=i, column=2, value=v if v is not None else "—").font = FONT_BODY

    _auto_width(ws)

    # ===== Hoja Cubicación =====
    ws_cub = wb.create_sheet("Cubicación")
    cols = ["#", "Partida", "Sistema", "Área", "Descripción", "Unidad",
            "Cantidad", "% Pérdida", "Cant. final", "PU (CLP)", "Total (CLP)",
            "Fuente", "Confianza", "Norma", "Observación"]
    _aplicar_header(ws_cub, 1, cols)

    total_materiales = 0
    for i, it in enumerate(cubicacion, start=1):
        cant = it.get("cantidad", 0) or 0
        factor = it.get("factor_perdida", 0) or 0
        pu = it.get("precio_unitario", 0) or 0
        cf = cant * (1 + factor / 100)
        total = cf * pu
        total_materiales += total

        fila = [
            i, it.get("partida"), it.get("sistema"), it.get("area"),
            it.get("descripcion"), it.get("unidad"),
            cant, factor, round(cf, 2), pu, round(total),
            it.get("fuente"), it.get("confianza"), it.get("norma_ref"),
            it.get("observacion"),
        ]
        for col_idx, val in enumerate(fila, start=1):
            c = ws_cub.cell(row=i + 1, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = BORDER_THIN
            if col_idx in (7, 8, 9, 10, 11):
                c.alignment = ALIGN_RIGHT
                if col_idx in (10, 11):
                    c.number_format = "#,##0"
            else:
                c.alignment = ALIGN_LEFT

    # Fila total
    fila_total = len(cubicacion) + 2
    ws_cub.cell(row=fila_total, column=10, value="TOTAL MATERIALES").font = FONT_TOTAL
    c = ws_cub.cell(row=fila_total, column=11, value=round(total_materiales))
    c.font = FONT_TOTAL
    c.number_format = "#,##0"
    c.alignment = ALIGN_RIGHT
    ws_cub.freeze_panes = "A2"
    _auto_width(ws_cub)

    # ===== Hoja Presupuesto =====
    ws_pres = wb.create_sheet("Presupuesto")
    cols_p = ["#", "Partida", "Descripción", "Unidad", "Cant. final",
              "PU mat.", "Total mat.", "HH/un", "HH tot.", "M.O. (CLP)", "Total ítem"]
    _aplicar_header(ws_pres, 1, cols_p)

    tarifa = (config or {}).get("tarifaHH", 0) or 0
    total_mat_p = 0
    total_mo_p = 0
    for i, it in enumerate(cubicacion, start=1):
        cant = it.get("cantidad", 0) or 0
        factor = it.get("factor_perdida", 0) or 0
        pu = it.get("precio_unitario", 0) or 0
        hhu = it.get("hh_unitaria", 0) or 0
        cf = cant * (1 + factor / 100)
        tot_mat = cf * pu
        hh_tot = cf * hhu
        mo_tot = hh_tot * tarifa
        tot_item = tot_mat + mo_tot
        total_mat_p += tot_mat
        total_mo_p += mo_tot
        fila = [
            i, it.get("partida"), it.get("descripcion"), it.get("unidad"),
            round(cf, 2), pu, round(tot_mat), hhu, round(hh_tot, 2),
            round(mo_tot), round(tot_item),
        ]
        for col_idx, val in enumerate(fila, start=1):
            c = ws_pres.cell(row=i + 1, column=col_idx, value=val)
            c.font = FONT_BODY
            c.border = BORDER_THIN
            if col_idx >= 5:
                c.alignment = ALIGN_RIGHT
                if col_idx in (6, 7, 10, 11):
                    c.number_format = "#,##0"

    # Resumen económico
    gg_pct = (config or {}).get("ggPct", 12)
    util_pct = (config or {}).get("utilidadPct", 10)
    iva_pct = (config or {}).get("ivaPct", 19)
    subtotal = total_mat_p + total_mo_p
    gg = subtotal * gg_pct / 100
    util = (subtotal + gg) * util_pct / 100
    neto = subtotal + gg + util
    iva = neto * iva_pct / 100
    total_final = neto + iva

    base_fila = len(cubicacion) + 3
    resumen = [
        ("Materiales", total_mat_p),
        ("Mano de obra", total_mo_p),
        ("Subtotal", subtotal),
        (f"GG ({gg_pct}%)", gg),
        (f"Utilidad ({util_pct}%)", util),
        ("Total neto", neto),
        (f"IVA ({iva_pct}%)", iva),
        ("TOTAL CON IVA", total_final),
    ]
    for i, (k, v) in enumerate(resumen):
        r = base_fila + i
        ws_pres.cell(row=r, column=10, value=k).font = FONT_TOTAL
        c = ws_pres.cell(row=r, column=11, value=round(v))
        c.font = FONT_TOTAL
        c.number_format = "#,##0"
        c.alignment = ALIGN_RIGHT

    ws_pres.freeze_panes = "A2"
    _auto_width(ws_pres)

    # ===== Hoja Normativa =====
    ws_n = wb.create_sheet("Normativa")
    cols_n = ["#", "Requisito", "Norma / referencia", "Artículo",
              "Estado", "Evidencia", "Observación"]
    _aplicar_header(ws_n, 1, cols_n)
    for i, n in enumerate(normativa, start=1):
        fila = [
            i, n.get("requisito"), n.get("norma_ref"), n.get("articulo"),
            n.get("estado"), n.get("evidencia"), n.get("observacion"),
        ]
        for col_idx, val in enumerate(fila, start=1):
            c = ws_n.cell(row=i + 1, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = BORDER_THIN
    ws_n.freeze_panes = "A2"
    _auto_width(ws_n)

    # ===== Hoja Riesgos =====
    ws_r = wb.create_sheet("Riesgos")
    cols_r = ["#", "Categoría", "Descripción", "Probabilidad", "Impacto",
              "Nivel", "Mitigación", "Responsable"]
    _aplicar_header(ws_r, 1, cols_r)
    for i, r in enumerate(riesgos, start=1):
        fila = [
            i, r.get("categoria"), r.get("descripcion"), r.get("probabilidad"),
            r.get("impacto"), r.get("nivel"), r.get("mitigacion"), r.get("responsable"),
        ]
        for col_idx, val in enumerate(fila, start=1):
            c = ws_r.cell(row=i + 1, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = BORDER_THIN
    ws_r.freeze_panes = "A2"
    _auto_width(ws_r)

    wb.save(str(output_path))
    return output_path
