"""
Parser de Excel (XLSX) usando openpyxl + pandas.
Detecta automáticamente columnas típicas de cubicación y normaliza el formato.
"""
from pathlib import Path
from typing import Any
import re

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    pd = None
    load_workbook = None


# Mapeo de columnas (claves canónicas y sus variantes en español)
COLUMNAS_CANONICAS = {
    "item":          ["item", "ítem", "n°", "no", "#", "nro", "número", "numero"],
    "descripcion":   ["descripción", "descripcion", "detalle", "concepto", "material"],
    "unidad":        ["unidad", "ud", "und", "u/m"],
    "cantidad":      ["cantidad", "cant", "qty", "cant."],
    "precio_unit":   ["precio unitario", "p.u.", "pu", "precio", "unit price", "precio unit"],
    "total":         ["total", "subtotal", "monto"],
    "partida":       ["partida", "capítulo", "capitulo", "rubro"],
    "area":          ["área", "area", "zona", "sector"],
    "sistema":       ["sistema", "system"],
    "norma":         ["norma", "estándar", "estandar"],
    "marca":         ["marca", "brand"],
    "modelo":        ["modelo", "model"],
    "proveedor":     ["proveedor", "supplier"],
    "observaciones": ["observaciones", "obs", "nota", "comentario"],
}


def _normalizar_columna(nombre: str) -> str | None:
    """Mapea un nombre de columna a su clave canónica."""
    if not nombre:
        return None
    n = re.sub(r"\s+", " ", str(nombre).strip().lower())
    for canon, variantes in COLUMNAS_CANONICAS.items():
        if n in variantes:
            return canon
    return None


def parse_xlsx(path: Path) -> dict[str, Any]:
    if pd is None or load_workbook is None:
        return {
            "tipo": "XLSX",
            "texto_plano": "",
            "estructura": {},
            "metadatos": {},
            "advertencias": ["pandas/openpyxl no instalados"],
        }

    hojas_info: list[dict] = []
    items_propuestos: list[dict] = []
    texto_total: list[str] = []
    advertencias: list[str] = []
    metadatos: dict[str, Any] = {}

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        metadatos["hojas_count"] = len(wb.sheetnames)
        metadatos["hojas"] = wb.sheetnames

        for hoja_nombre in wb.sheetnames:
            # Leer con pandas para procesamiento más robusto
            try:
                df = pd.read_excel(path, sheet_name=hoja_nombre, header=None)
            except Exception as e:
                advertencias.append(f"No se pudo leer hoja '{hoja_nombre}': {e}")
                continue

            if df.empty:
                hojas_info.append({"hoja": hoja_nombre, "filas": 0, "cols": 0, "advertencia": "vacía"})
                continue

            # Detectar fila de cabecera: la primera fila donde se reconozcan ≥ 2 columnas canónicas
            header_row = None
            mapa_columnas = {}
            for i in range(min(15, len(df))):
                fila = df.iloc[i]
                mapa = {}
                for col_idx, val in enumerate(fila):
                    canon = _normalizar_columna(val)
                    if canon:
                        mapa[col_idx] = canon
                if len(mapa) >= 2:
                    header_row = i
                    mapa_columnas = mapa
                    break

            hoja_info = {
                "hoja": hoja_nombre,
                "filas": int(len(df)),
                "cols": int(df.shape[1]),
                "header_row": header_row,
                "columnas_detectadas": list(mapa_columnas.values()),
            }
            hojas_info.append(hoja_info)

            # Si detectamos cabeceras, extraer items
            if header_row is not None:
                datos = df.iloc[header_row + 1:].reset_index(drop=True)
                for _, fila in datos.iterrows():
                    item = {}
                    for col_idx, canon in mapa_columnas.items():
                        try:
                            valor = fila.iloc[col_idx]
                        except IndexError:
                            continue
                        if pd.isna(valor):
                            continue
                        item[canon] = str(valor).strip() if isinstance(valor, str) else valor

                    # Solo agregar si tiene al menos descripción + (cantidad o unidad)
                    if item.get("descripcion") and (item.get("cantidad") or item.get("unidad")):
                        item["_hoja"] = hoja_nombre
                        items_propuestos.append(item)

            # Concatenar texto para RAG
            try:
                texto_total.append(df.to_string(index=False, header=False, na_rep=""))
            except Exception:
                pass

        wb.close()

    except Exception as e:
        return {
            "tipo": "XLSX",
            "texto_plano": "",
            "estructura": {},
            "metadatos": {},
            "advertencias": [f"Error procesando XLSX: {e}"],
        }

    return {
        "tipo": "XLSX",
        "texto_plano": "\n\n".join(texto_total),
        "estructura": {
            "hojas": hojas_info,
            "items_detectados": len(items_propuestos),
            "items": items_propuestos[:500],  # cap para JSON
        },
        "metadatos": metadatos,
        "advertencias": advertencias,
    }
